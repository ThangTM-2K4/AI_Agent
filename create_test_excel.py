from openpyxl import Workbook
from datetime import datetime

# Create a test Excel file with unlearned courses (grade X)
wb = Workbook()
ws = wb.active
ws.title = "điểm"

# Add header row with student info
ws['A1'] = "Mã SV/Mã HV: SV202601 - Họ tên: Test Student - Lớp: CNTT-K18"

# Add empty row
ws['A2'] = ""

# Add column headers
headers = ["Học phần", "Tín chỉ", "Điểm Chữ", "D1", "D2", "Năm học", "Kỳ"]
for col, header in enumerate(headers, 1):
    ws.cell(row=4, column=col, value=header)

# Add some courses with grades
courses = [
    ("Toán", 3, "A", 8.5, 8.5, "2024", "Kỳ 1"),
    ("Lập trình", 4, "B", 7.0, 7.0, "2024", "Kỳ 1"),
    ("Cơ sở dữ liệu", 3, "X", None, None, "2024", "Kỳ 2"),  # Chưa học
    ("Phát triển web", 3, "X", None, None, "2024", "Kỳ 2"),  # Chưa học
    ("Mạng máy tính", 4, "C", 5.5, 5.5, "2024", "Kỳ 2"),
]

for row_idx, (name, credits, grade, d1, d2, year, ky) in enumerate(courses, 5):
    ws.cell(row=row_idx, column=1, value=name)
    ws.cell(row=row_idx, column=2, value=credits)
    ws.cell(row=row_idx, column=3, value=grade)
    ws.cell(row=row_idx, column=4, value=d1)
    ws.cell(row=row_idx, column=5, value=d2)
    ws.cell(row=row_idx, column=6, value=year)
    ws.cell(row=row_idx, column=7, value=ky)

# Save file
wb.save("test_with_x.xlsx")
print("Created test_with_x.xlsx with 2 unlearned courses (grade X)")

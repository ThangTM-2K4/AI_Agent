from datetime import date, datetime
import re
from io import BytesIO
from typing import Any
import unicodedata

from openpyxl import load_workbook
from sqlalchemy import delete, select
from sqlalchemy.orm import Session

from app.models.student import CourseResult, LearningActivity, Student
from app.services.analytics_service import AnalyticsService


VALID_GRADES = {"A", "B", "C", "D", "F", "X"}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("đ", "d")
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_")


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sheet_to_dicts(sheet) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_header(header) for header in rows[0]]
    records: list[dict[str, Any]] = []

    for row in rows[1:]:
        if not row or all(cell is None for cell in row):
            continue
        record = {}
        for index, header in enumerate(headers):
            if header:
                record[header] = row[index] if index < len(row) else None
        records.append(record)

    return records


def _parse_student_meta_from_title(title_text: str) -> tuple[str, str, str | None, str | None]:
    text = _to_text(title_text)

    code_match = re.search(r"mã\s*(?:sv|hv)\s*:\s*([\w\-/.]+)", text, flags=re.IGNORECASE)
    name_match = re.search(
        r"họ\s*tên\s*:\s*(.*?)(?:\s*-\s*lớp\s*:|\s*-\s*trạng\s*thái\s*:|\s*\(|$)",
        text,
        flags=re.IGNORECASE,
    )
    class_match = re.search(
        r"lớp\s*:\s*(.*?)(?:\s*-\s*trạng\s*thái\s*:|\s*\(|$)",
        text,
        flags=re.IGNORECASE,
    )
    status_match = re.search(r"trạng\s*thái\s*:\s*(.*?)(?:\s*\(|$)", text, flags=re.IGNORECASE)

    student_code = code_match.group(1).strip() if code_match else ""
    full_name = name_match.group(1).strip() if name_match else ""
    class_name = class_match.group(1).strip() if class_match else None
    status = status_match.group(1).strip() if status_match else None

    if not student_code or not full_name:
        raise ValueError(
            "Không đọc được mã sinh viên/họ tên từ dòng tiêu đề. Vui lòng giữ dạng 'Mã SV/Mã HV: ... - Họ tên: ...'."
        )

    return student_code, full_name, class_name, status


def _find_meta_text(raw_rows: list[tuple[Any, ...]]) -> str:
    for row in raw_rows[:8]:
        for cell in row:
            text = _to_text(cell)
            if not text:
                continue
            lowered = text.lower()
            if "mã sv" in lowered or "ma sv" in lowered or "họ tên" in lowered or "ho ten" in lowered:
                return text
    return ""


def _find_header_row_index(raw_rows: list[tuple[Any, ...]]) -> int:
    for idx, row in enumerate(raw_rows[:20]):
        headers = {_normalize_header(cell) for cell in row if _to_text(cell)}
        if "hoc_phan" in headers and ("tchi" in headers or "d1" in headers or "dbp" in headers):
            return idx
    raise ValueError("Không tìm thấy dòng tiêu đề cột trong sheet 'điểm'.")


def _parse_date(value: Any) -> date:
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        return datetime.strptime(value.strip(), "%Y-%m-%d").date()
    raise ValueError(f"Không thể parse ngày: {value}")


def _to_float_or_none(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)

    text = _to_text(value)
    if not text:
        return None

    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _to_int_or_none(value: Any) -> int | None:
    number = _to_float_or_none(value)
    if number is None:
        return None

    # TChi is expected to be an integer field; reject non-integer values.
    if not float(number).is_integer():
        return None

    return int(number)


def _extract_score_d1_d2(row: tuple[Any, ...], header_map: dict[str, int]) -> tuple[float, float | None, float | None]:
    """
    Lấy D1, D2 từ row, return (score_để_dùng, score_d1, score_d2).
    
    score_để_dùng: D2 nếu có, nếu không thì D1 (logic cũ để backward compatible).
    score_d1: giá trị D1 gốc (có thể None).
    score_d2: giá trị D2 gốc (có thể None).
    """
    score_d1_val = _normalize_score_value(_row_value(row, header_map, "d1", "d_1"))
    score_d2_val = _normalize_score_value(_row_value(row, header_map, "d2", "d_2"))

    # D2 có giá trị thì dùng D2, nếu D2 trống mới fallback sang D1.
    score = score_d2_val if score_d2_val is not None else score_d1_val

    return score if score is not None else 0.0, score_d1_val, score_d2_val

    
def _normalize_text_for_match(value: Any) -> str:
    text = _to_text(value).lower()
    text = unicodedata.normalize("NFD", text)
    text = "".join(char for char in text if unicodedata.category(char) != "Mn")
    text = text.replace("đ", "d")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _is_excluded_course_name(name: Any) -> bool:
    normalized_name = _normalize_text_for_match(name)
    return any(keyword in normalized_name for keyword in ("thuc tap", "khoa luan"))


def _is_student_meta_line(row: tuple[Any, ...]) -> bool:
    row_text = " ".join(_to_text(cell) for cell in row if _to_text(cell)).lower()
    return (
        ("mã sv" in row_text or "ma sv" in row_text or "mã hv" in row_text or "ma hv" in row_text)
        and ("họ tên" in row_text or "ho ten" in row_text or "họ và tên" in row_text or "ho va ten" in row_text)
    )


def _header_map_from_row(row: tuple[Any, ...]) -> dict[str, int]:
    return {_normalize_header(cell): idx for idx, cell in enumerate(row) if _to_text(cell)}


def _row_value(row: tuple[Any, ...], header_map: dict[str, int], *names: str) -> Any:
    for name in names:
        idx = header_map.get(name)
        if idx is not None and idx < len(row):
            value = row[idx]
            if value is not None and value != "":
                return value
    return None


def _normalize_score_value(value: Any) -> float | None:
    return AnalyticsService.normalize_score_value(value)


def _extract_blocked_score_sheet_students(raw_rows: list[tuple[Any, ...]]) -> dict[str, dict[str, Any]]:
    students: dict[str, dict[str, Any]] = {}
    current_code: str | None = None
    current_headers: dict[str, int] | None = None

    for row in raw_rows:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        if _is_student_meta_line(row):
            meta_text = " ".join(_to_text(cell) for cell in row if _to_text(cell))
            try:
                student_code, full_name, class_name, status = _parse_student_meta_from_title(meta_text)
            except ValueError:
                current_code = None
                current_headers = None
                continue

            current_code = student_code
            current_headers = None
            if student_code not in students:
                students[student_code] = {
                    "full_name": full_name,
                    "class_name": class_name,
                    "status": status,
                    "subjects": [],
                    "seen": set(),
                }
            else:
                students[student_code]["full_name"] = full_name
                students[student_code]["class_name"] = class_name
                students[student_code]["status"] = status
            continue

        if current_code is None:
            continue

        normalized_headers = {_normalize_header(cell) for cell in row if _to_text(cell)}
        if "hoc_phan" in normalized_headers and ("tchi" in normalized_headers or "d1" in normalized_headers or "dbp" in normalized_headers):
            current_headers = _header_map_from_row(row)
            continue

        if current_headers is None:
            continue

        course_name = _to_text(_row_value(row, current_headers, "hoc_phan", "mon_hoc", "course_name"))
        if not course_name:
            continue

        grade = _to_text(_row_value(row, current_headers, "dchu", "d_chu", "diem_chu")).upper()
        if grade not in VALID_GRADES:
            continue
        if _is_excluded_course_name(course_name) and grade != "X":
            continue

        score, score_d1_val, score_d2_val = _extract_score_d1_d2(row, current_headers)
        if score is None and grade != "X":
            continue

        credit = _to_int_or_none(_row_value(row, current_headers, "tchi", "t_chi", "tin_chi", "credits"))
        if (credit is None or credit <= 0) and grade != "X":
            continue
        if credit is None or credit <= 0:
            credit = 0

        subject_key = (course_name, _row_value(row, current_headers, "nam_hoc", "academic_year"), _row_value(row, current_headers, "ky", "semester"))
        seen = students[current_code]["seen"]
        if subject_key in seen:
            continue
        seen.add(subject_key)

        students[current_code]["subjects"].append(
            {
                "name": course_name,
                "score": float(score),
                "score_d1": score_d1_val,
                "score_d2": score_d2_val,
                "credit": credit,
                "year": _to_text(_row_value(row, current_headers, "nam_hoc", "academic_year")),
                "semester": _to_text(_row_value(row, current_headers, "ky", "semester")),
                "grade": grade,
            }
        )

    for student_data in students.values():
        student_data.pop("seen", None)

    return students


def _extract_score_sheet_subjects(raw_rows: list[tuple[Any, ...]]) -> list[dict[str, Any]]:
    header_index = _find_header_row_index(raw_rows)
    header_row = [_normalize_header(cell) for cell in raw_rows[header_index]]
    column_index = {header: idx for idx, header in enumerate(header_row) if header}

    def _value(row: tuple[Any, ...], *names: str) -> Any:
        for name in names:
            index = column_index.get(name)
            if index is not None and index < len(row):
                value = row[index]
                if value is not None and value != "":
                    return value
        return None

    subjects: list[dict[str, Any]] = []
    seen_courses: set[str] = set()

    for row in raw_rows[header_index + 1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        name = _to_text(_value(row, "hoc_phan", "mon_hoc", "course_name"))
        if not name or len(name) < 3:
            continue

        grade = _to_text(_value(row, "dchu", "d_chu", "diem_chu")).upper()
        if grade not in VALID_GRADES:
            continue
        if _is_excluded_course_name(name) and grade != "X":
            continue

        score, score_d1_val, score_d2_val = _extract_score_d1_d2(row, column_index)
        if score is None and grade != "X":
            continue

        credit = _to_int_or_none(_value(row, "tchi", "t_chi", "tin_chi", "credits"))
        if (credit is None or credit <= 0) and grade != "X":
            continue
        if credit is None or credit <= 0:
            credit = 0

        if name in seen_courses:
            continue
        seen_courses.add(name)

        subjects.append(
            {
                "name": name,
                "score": float(score),
                "score_d1": score_d1_val,
                "score_d2": score_d2_val,
                "credit": credit,
                "year": _to_text(_value(row, "nam_hoc", "academic_year")),
                "semester": _to_text(_value(row, "ky", "semester")),
                "grade": grade,
            }
        )

    return subjects


def _extract_multi_student_score_sheet(raw_rows: list[tuple[Any, ...]]) -> dict[str, dict[str, Any]]:
    header_index = _find_header_row_index(raw_rows)
    header_row = [_normalize_header(cell) for cell in raw_rows[header_index]]
    column_index = {header: idx for idx, header in enumerate(header_row) if header}

    student_code_columns = ("student_code", "mssv", "ma_sv", "ma_hv", "ma_sv_ma_hv")
    if not any(name in column_index for name in student_code_columns):
        return {}

    def _value(row: tuple[Any, ...], *names: str) -> Any:
        for name in names:
            index = column_index.get(name)
            if index is not None and index < len(row):
                value = row[index]
                if value is not None and value != "":
                    return value
        return None

    by_student: dict[str, dict[str, Any]] = {}

    for row in raw_rows[header_index + 1:]:
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        student_code = _to_text(_value(row, "student_code", "mssv", "ma_sv", "ma_hv", "ma_sv_ma_hv"))
        full_name = _to_text(_value(row, "full_name", "ho_ten", "hoten", "ten_sinh_vien", "ho_va_ten"))
        class_name = _to_text(_value(row, "class_name", "lop")) or None
        if not student_code or not full_name:
            continue

        course_name = _to_text(_value(row, "hoc_phan", "mon_hoc", "course_name"))
        if not course_name or len(course_name) < 3:
            continue

        grade = _to_text(_value(row, "dchu", "d_chu", "diem_chu")).upper()
        if grade not in VALID_GRADES:
            continue
        if _is_excluded_course_name(course_name) and grade != "X":
            continue

        score, score_d1_val, score_d2_val = _extract_score_d1_d2(row, column_index)
        if score is None and grade != "X":
            continue

        credit = _to_int_or_none(_value(row, "tchi", "t_chi", "tin_chi", "credits"))
        if (credit is None or credit <= 0) and grade != "X":
            continue
        if credit is None or credit <= 0:
            credit = 0

        year = _to_text(_value(row, "nam_hoc", "academic_year"))
        semester = _to_text(_value(row, "ky", "semester"))

        if student_code not in by_student:
            by_student[student_code] = {
                "full_name": full_name,
                "class_name": class_name,
                "subjects": [],
                "seen": set(),
            }

        key = (course_name, year, semester)
        if key in by_student[student_code]["seen"]:
            continue
        by_student[student_code]["seen"].add(key)

        by_student[student_code]["subjects"].append(
            {
                "name": course_name,
                "score": float(score),
                "score_d1": score_d1_val,
                "score_d2": score_d2_val,
                "credit": credit,
                "year": year,
                "semester": semester,
                "grade": grade,
            }
        )

    # remove helper set before returning
    for student_data in by_student.values():
        student_data.pop("seen", None)

    return by_student



class ExcelImportService:
    def __init__(self, db: Session):
        self.db = db

    @staticmethod
    def _find_sheet(workbook, candidates: list[str]):
        normalized_candidates = {candidate.lower(): candidate for candidate in candidates}
        for sheet_name in workbook.sheetnames:
            key = sheet_name.strip().lower()
            if key in normalized_candidates:
                return workbook[sheet_name]
        return None

    @staticmethod
    def _find_single_score_sheet(workbook):
        for sheet_name in workbook.sheetnames:
            lowered = sheet_name.strip().lower()
            if lowered in {"điểm", "diem", "bang diem", "bảng điểm"}:
                return workbook[sheet_name]
        return None

    def _upsert_student(self, student_code: str, full_name: str, class_name: str | None) -> Student:
        student = self.db.execute(
            select(Student).where(Student.student_code == student_code)
        ).scalars().first()

        if student is None:
            student = Student(
                student_code=student_code,
                full_name=full_name,
                class_name=class_name,
            )
            self.db.add(student)
        else:
            student.full_name = full_name
            student.class_name = class_name

        self.db.flush()
        return student

    def import_workbook(self, file_bytes: bytes) -> dict[str, int | str]:
        workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True)
        print("debug", workbook)

        students_sheet = self._find_sheet(workbook, ["students", "student", "sinhvien", "sinh_vien"])
        results_sheet = self._find_sheet(workbook, ["course_results", "results", "diem", "diem_so"])
        activities_sheet = self._find_sheet(workbook, ["learning_activities", "activities", "hanh_vi_hoc_tap"])
        score_sheet = self._find_single_score_sheet(workbook)

        if students_sheet is None and score_sheet is None:
            raise ValueError(
                "File Excel phải có sheet Students hoặc sheet 'điểm'. Hiện tại file của bạn có vẻ là workbook điểm theo từng sinh viên."
            )

        imported_students = 0
        imported_results = 0
        imported_activities = 0
        student_map: dict[str, Student] = {}

        if score_sheet is not None:
            raw_rows = list(score_sheet.iter_rows(values_only=True))
            if len(raw_rows) < 3:
                raise ValueError("Sheet 'điểm' không có đủ dữ liệu.")

            multi_student_data = _extract_blocked_score_sheet_students(raw_rows)
            if not multi_student_data:
                multi_student_data = _extract_multi_student_score_sheet(raw_rows)
            if multi_student_data:
                student_ids_to_replace: set[int] = set()
                result_rows: list[CourseResult] = []

                for student_code, data in multi_student_data.items():
                    subjects = data["subjects"]
                    if not subjects:
                        continue

                    student = self._upsert_student(
                        student_code,
                        _to_text(data.get("full_name")),
                        _to_text(data.get("class_name")) or None,
                    )
                    student_map[student_code] = student
                    imported_students += 1
                    student_ids_to_replace.add(student.id)

                    for subject in subjects:
                        semester = subject.get("semester") or "UNKNOWN"
                        year = subject.get("year") or ""
                        if year and semester:
                            semester = f"{year}-{semester}"
                        elif year:
                            semester = year

                        course_name = subject["name"]
                        course_code = _to_text(course_name).upper().replace(" ", "_")[:20]
                        result_rows.append(
                            CourseResult(
                                student_id=student.id,
                                semester=semester or "UNKNOWN",
                                course_code=course_code,
                                course_name=course_name,
                                credits=int(subject["credit"]),
                                score=float(subject["score"]),
                                score_d1=subject.get("score_d1"),
                                score_d2=subject.get("score_d2"),
                                grade=subject.get("grade"),
                            )
                        )

                if not result_rows:
                    raise ValueError(
                        "Không tìm thấy môn hợp lệ trong sheet 'điểm'. Chỉ lấy A/B/C/D/F, bỏ X/P và bỏ dòng thiếu điểm hoặc tín chỉ."
                    )

                if student_ids_to_replace:
                    self.db.execute(delete(CourseResult).where(CourseResult.student_id.in_(student_ids_to_replace)))
                self.db.add_all(result_rows)
                imported_results = len(result_rows)
            else:
                meta_text = _find_meta_text(raw_rows)
                student_code, full_name, class_name, _status = _parse_student_meta_from_title(meta_text)
                student = self._upsert_student(student_code, full_name, class_name)
                student_map[student_code] = student
                imported_students += 1

                subjects = _extract_score_sheet_subjects(raw_rows)
                if not subjects:
                    raise ValueError("Không tìm thấy môn hợp lệ trong sheet 'điểm'.")

                result_rows: list[CourseResult] = []
                student_ids_to_replace = {student.id}

                for subject in subjects:
                    semester = subject.get("semester") or "UNKNOWN"
                    year = subject.get("year") or ""
                    if year and semester:
                        semester = f"{year}-{semester}"
                    elif year:
                        semester = year

                    course_name = subject["name"]
                    course_code = _to_text(course_name).upper().replace(" ", "_")[:20]

                    result_rows.append(
                        CourseResult(
                            student_id=student.id,
                            semester=semester or "UNKNOWN",
                            course_code=course_code,
                            course_name=course_name,
                            credits=int(subject["credit"]),
                            score=float(subject["score"]),
                            score_d1=subject.get("score_d1"),
                            score_d2=subject.get("score_d2"),
                            grade=subject.get("grade"),
                        )
                    )

                if student_ids_to_replace:
                    self.db.execute(delete(CourseResult).where(CourseResult.student_id.in_(student_ids_to_replace)))
                self.db.add_all(result_rows)
                imported_results = len(result_rows)

        if students_sheet is not None:
            student_rows = _sheet_to_dicts(students_sheet)
            if not student_rows:
                raise ValueError("Sheet Students không có dữ liệu.")

            for row in student_rows:
                student_code = _to_text(row.get("student_code") or row.get("mssv") or row.get("ma_sv") or row.get("ma_sv/mã_hv"))
                full_name = _to_text(row.get("full_name") or row.get("ho_ten") or row.get("hoten"))
                class_name = row.get("class_name") or row.get("lop")

                if not student_code or not full_name:
                    raise ValueError("Mỗi dòng trong sheet Students phải có student_code và full_name.")

                student = self._upsert_student(student_code, full_name, _to_text(class_name) or None)
                student_map[student_code] = student
                imported_students += 1

        if score_sheet is None and results_sheet is not None:
            result_rows = _sheet_to_dicts(results_sheet)
            if result_rows:
                student_ids_to_replace: set[int] = set()
                prepared_results: list[CourseResult] = []

                for row in result_rows:
                    student_code = _to_text(row.get("student_code") or row.get("mssv"))
                    if student_code not in student_map:
                        raise ValueError(
                            f"Course results tham chiếu student_code không tồn tại trong sheet Students: {student_code}"
                        )

                    grade = _to_text(row.get("grade") or row.get("dchu") or row.get("diem_chu")).upper()
                    if grade and grade not in VALID_GRADES:
                        continue

                    course_name = _to_text(row.get("course_name") or row.get("hoc_phan") or row.get("mon_hoc"))
                    if not course_name:
                        continue
                    if _is_excluded_course_name(course_name) and grade != "X":
                        continue

                    score_d1_val = _normalize_score_value(row.get("d1"))
                    score_d2_val = _normalize_score_value(row.get("d2") if row.get("d2") is not None else row.get("score"))
                    score = score_d2_val if score_d2_val is not None else score_d1_val
                    
                    credit = _to_int_or_none(row.get("credits") or row.get("tchi") or row.get("tin_chi"))
                    if score is None or ((credit is None or credit <= 0) and grade != "X"):
                        continue
                    if credit is None or credit <= 0:
                        credit = 0

                    student = student_map[student_code]
                    student_ids_to_replace.add(student.id)
                    prepared_results.append(
                        CourseResult(
                            student_id=student.id,
                            semester=_to_text(row.get("semester")) or "UNKNOWN",
                            course_code=_to_text(row.get("course_code"))[:20],
                            course_name=course_name,
                            credits=credit,
                            score=score,
                            score_d1=score_d1_val,
                            score_d2=score_d2_val,
                            grade=grade,
                        )
                    )

                if student_ids_to_replace:
                    self.db.execute(delete(CourseResult).where(CourseResult.student_id.in_(student_ids_to_replace)))
                self.db.add_all(prepared_results)
                imported_results = len(prepared_results)

        if activities_sheet is not None:
            activity_rows = _sheet_to_dicts(activities_sheet)
            if activity_rows:
                student_ids_to_replace: set[int] = set()
                prepared_activities: list[LearningActivity] = []

                for row in activity_rows:
                    student_code = _to_text(row.get("student_code") or row.get("mssv"))
                    if student_code not in student_map:
                        raise ValueError(
                            f"Learning activities tham chiếu student_code không tồn tại trong sheet Students: {student_code}"
                        )

                    student = student_map[student_code]
                    student_ids_to_replace.add(student.id)
                    prepared_activities.append(
                        LearningActivity(
                            student_id=student.id,
                            event_date=_parse_date(row.get("event_date") or row.get("ngay")),
                            metric_name=_to_text(row.get("metric_name") or row.get("ten_chi_so")),
                            metric_value=float(row.get("metric_value") or 0),
                        )
                    )

                if student_ids_to_replace:
                    self.db.execute(delete(LearningActivity).where(LearningActivity.student_id.in_(student_ids_to_replace)))
                self.db.add_all(prepared_activities)
                imported_activities = len(prepared_activities)

        self.db.commit()
        return {
            "students_upserted": imported_students,
            "course_results_imported": imported_results,
            "learning_activities_imported": imported_activities,
            "message": "Import Excel thành công.",
        }